from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator

from .action_tracker import ACTIONED_COLUMNS, build_ticket_action_views
from .models import AssignmentTicket, CompletionTicket, FieldActivityReport
from .ticket_dedupe import dedupe_assignments, dedupe_completions

CONNECTION_COLUMNS = [
    "ACC",
    "NAME",
    "LOCATION",
    "CONTACT",
    "CLOSE DATE/TIME",
    "SN",
    "ONU",
    "INDOOR CABLE",
    "OUTDOOR CABLE",
    "ATB",
    "PATCH",
    "POWER LEVEL",
    "POWER FAT",
    "TECH",
    "REMARKS",
]
ASSIGNMENT_COLUMNS = [
    "ACC",
    "NAME",
    "CONTACT",
    "ROUTE",
    "LOCATION",
    "TECH",
    "STATUS",
    "REMARKS",
]
FIELD_ACTIVITY_COLUMNS = [
    "CONTRACTOR",
    "LOCATION",
    "SCOPE",
    "POB",
    "TOPIC",
]
SUMMARY_COUNT_COLUMNS = ["ITEM", "COUNT"]


def ensure_output_from_template(template_path: str | Path, output_path: str | Path) -> Path:
    template = Path(template_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if template.resolve() != output.resolve():
        shutil.copy2(template, output)
    return output


def next_empty_row(ws) -> int:
    for row in range(ws.max_row, 1, -1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            return row + 1
    return 2


def copy_formula_if_present(ws, source_row: int, target_row: int, column: int) -> None:
    source_cell = ws.cell(source_row, column)
    target_cell = ws.cell(target_row, column)
    if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
        translated = Translator(source_cell.value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
        target_cell.value = translated


def append_connections(
    template_path: str | Path,
    output_path: str | Path,
    records: list[CompletionTicket],
    sheet_name: str | None = None,
) -> Path:
    records = dedupe_completions(records)
    output = ensure_output_from_template(template_path, output_path)
    wb = load_workbook(output)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}

    row_pointer = next_empty_row(ws)
    source_formula_row = max(2, row_pointer - 1)
    for record in records:
        row_values = record.to_row()
        for name in CONNECTION_COLUMNS:
            column = header_to_col.get(name)
            if not column:
                continue
            value = row_values.get(name)
            if value not in (None, ""):
                ws.cell(row_pointer, column).value = value
            else:
                copy_formula_if_present(ws, source_formula_row, row_pointer, column)
        row_pointer += 1

    wb.save(output)
    return output


def assignment_rows(records: list[AssignmentTicket]) -> list[list[str | None]]:
    rows: list[list[str | None]] = [ASSIGNMENT_COLUMNS]
    for record in dedupe_assignments(records):
        rows.append(
            [
                record.account,
                record.client_name,
                record.contact,
                record.route_code,
                record.location,
                record.tech,
                record.status,
                record.remarks,
            ]
        )
    return rows


def field_activity_rows(records: list[FieldActivityReport]) -> list[list[str | int | None]]:
    rows: list[list[str | int | None]] = [FIELD_ACTIVITY_COLUMNS]
    for record in records:
        rows.append(
            [
                record.contractor,
                record.location,
                record.scope,
                record.pob,
                record.topic,
            ]
        )
    return rows


def summary_count_rows(summary_counts: dict[str, int]) -> list[list[str | int]]:
    rows: list[list[str | int]] = [SUMMARY_COUNT_COLUMNS]
    for key, value in summary_counts.items():
        rows.append([key, value])
    return rows


def ticket_action_rows(rows_data: list[dict[str, str | None]]) -> list[list[str | None]]:
    rows: list[list[str | None]] = [ACTIONED_COLUMNS]
    for record in rows_data:
        rows.append([record.get(column) for column in ACTIONED_COLUMNS])
    return rows


def write_rows_to_sheet(ws, rows: list[list[str | int | float | None]]) -> None:
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value


def write_assignment_sheet(
    workbook_path: str | Path,
    records: list[AssignmentTicket],
    sheet_name: str = "Assignments",
) -> Path:
    workbook = Path(workbook_path)
    wb = load_workbook(workbook)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)
    write_rows_to_sheet(ws, assignment_rows(records))
    wb.save(workbook)
    return workbook


def create_parsed_report_workbook(
    output_path: str | Path,
    assignments: list[AssignmentTicket],
    field_activity_reports: list[FieldActivityReport],
    completions: list[CompletionTicket],
    summary_counts: dict[str, int],
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    actioned_tickets, not_actioned_tickets = build_ticket_action_views(assignments, completions)

    wb = Workbook()
    ws_assignments = wb.active
    ws_assignments.title = "Parsed Assignments"
    write_rows_to_sheet(ws_assignments, assignment_rows(assignments))

    ws_field = wb.create_sheet(title="Parsed Field Activity")
    write_rows_to_sheet(ws_field, field_activity_rows(field_activity_reports))

    ws_completions = wb.create_sheet(title="Parsed Completions")
    completion_sheet_rows: list[list[str | int | float | None]] = [CONNECTION_COLUMNS]
    for record in dedupe_completions(completions):
        row = record.to_row()
        completion_sheet_rows.append([row.get(header) for header in CONNECTION_COLUMNS])
    write_rows_to_sheet(ws_completions, completion_sheet_rows)

    ws_summary = wb.create_sheet(title="Summary Counts")
    write_rows_to_sheet(ws_summary, summary_count_rows(summary_counts))

    ws_actioned = wb.create_sheet(title="Actioned Tickets")
    write_rows_to_sheet(ws_actioned, ticket_action_rows(actioned_tickets))

    ws_not_actioned = wb.create_sheet(title="Not Actioned Tickets")
    write_rows_to_sheet(ws_not_actioned, ticket_action_rows(not_actioned_tickets))

    wb.save(output)
    return output


def find_or_create_date_column(ws, target_date: date) -> int:
    for col in range(3, ws.max_column + 1):
        value = ws.cell(1, col).value
        if isinstance(value, datetime):
            if value.date() == target_date:
                return col
        elif isinstance(value, date) and value == target_date:
            return col

    total_col = ws.max_column
    if str(ws.cell(1, total_col).value).strip().lower() == "total":
        insert_col = total_col
        ws.insert_cols(insert_col)
        ws.cell(1, insert_col).value = datetime.combine(target_date, datetime.min.time())
        return insert_col

    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = datetime.combine(target_date, datetime.min.time())
    return new_col


def update_summary(
    template_path: str | Path,
    output_path: str | Path,
    summary_counts: dict[str, int],
    target_date: date,
    sheet_name: str | None = None,
) -> Path:
    output = ensure_output_from_template(template_path, output_path)
    wb = load_workbook(output)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    date_col = find_or_create_date_column(ws, target_date)

    row_map = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 2).value
        if label:
            row_map[str(label).strip()] = row

    for label, value in summary_counts.items():
        row = row_map.get(label)
        if row:
            value_cell = ws.cell(row, date_col)
            value_cell.value = value
            value_cell.number_format = "0"
            total_col = ws.max_column
            if str(ws.cell(1, total_col).value).strip().lower() == "total":
                start_letter = ws.cell(1, 3).column_letter
                end_letter = ws.cell(1, total_col - 1).column_letter
                ws.cell(row, total_col).value = f"=SUM({start_letter}{row}:{end_letter}{row})"

    wb.save(output)
    return output
