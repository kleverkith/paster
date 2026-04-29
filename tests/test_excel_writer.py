from datetime import date

from openpyxl import Workbook, load_workbook

from paster.excel_writer import append_connections, create_parsed_report_workbook, write_assignment_sheet
from paster.models import AssignmentTicket, CompletionTicket, FieldActivityReport


def build_connections_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "April Connections"
    headers = [
        "ACC",
        "NAME",
        "LOCATION",
        "CONTACT",
        "CLOSE DATE/TIME",
        "SN",
        "ONU",
        "INDOOR CABLE",
        "OUTDOOR CABLE",
        "ATB",
        "PATCH",
        "POWER LEVEL",
        "POWER FAT",
        "TECH",
        "REMARKS",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(1, col).value = value
    wb.save(path)


def test_write_assignment_sheet_exports_status_and_remarks(tmp_path):
    template = tmp_path / "sample.xlsx"
    output = tmp_path / "output.xlsx"
    build_connections_template(template)

    append_connections(
        template_path=template,
        output_path=output,
        records=[CompletionTicket(account="SFKL1", client_name="Done Client")],
        sheet_name="April Connections",
    )
    write_assignment_sheet(
        workbook_path=output,
        records=[
            AssignmentTicket(
                account="SFKL2",
                client_name="Waiting Client",
                contact="0712345678",
                route_code="02-S-FKLN-01",
                location="KILIMANI ROAD",
                status="Assigned",
                remarks="Scheduled tomorrow",
            )
        ],
        sheet_name="Assignments",
    )

    wb = load_workbook(output)
    ws = wb["Assignments"]
    assert ws.cell(1, 1).value == "ACC"
    assert ws.cell(1, 6).value == "TECH"
    assert ws.cell(1, 7).value == "STATUS"
    assert ws.cell(1, 8).value == "REMARKS"
    assert ws.cell(2, 1).value == "SFKL2"
    assert ws.cell(2, 5).value == "KILIMANI ROAD"
    assert ws.cell(2, 6).value is None
    assert ws.cell(2, 7).value == "Assigned"
    assert ws.cell(2, 8).value == "Scheduled tomorrow"


def test_create_parsed_report_workbook_creates_four_expected_sheets(tmp_path):
    output = tmp_path / "parsed-report.xlsx"

    create_parsed_report_workbook(
        output_path=output,
        assignments=[
            AssignmentTicket(account="SFKL2", client_name="Waiting Client", status="Assigned", remarks="Scheduled tomorrow")
        ],
        field_activity_reports=[
            FieldActivityReport(contractor="Comcraft", location="langata", scope="FTTH", pob=2, topic="severe weather")
        ],
        completions=[
            CompletionTicket(account="SFKL1", client_name="Done Client", location="Delamere Flats", remarks="Connected")
        ],
        summary_counts={"Assigned Tickets": 1, "Tickets done": 1, "Teams Available": 1},
    )

    wb = load_workbook(output)
    assert wb.sheetnames == [
        "Parsed Assignments",
        "Parsed Field Activity",
        "Parsed Completions",
        "Summary Counts",
        "Actioned Tickets",
        "Not Actioned Tickets",
    ]
    assert wb["Parsed Assignments"].cell(1, 1).value == "ACC"
    assert wb["Parsed Assignments"].cell(2, 1).value == "SFKL2"
    assert wb["Parsed Field Activity"].cell(1, 1).value == "CONTRACTOR"
    assert wb["Parsed Field Activity"].cell(2, 4).value == 2
    assert wb["Parsed Completions"].cell(1, 1).value == "ACC"
    assert wb["Parsed Completions"].cell(2, 1).value == "SFKL1"
    assert wb["Parsed Completions"].cell(2, 15).value == "Connected"
    assert wb["Actioned Tickets"].cell(1, 1).value == "ACC"
    assert wb["Not Actioned Tickets"].cell(1, 1).value == "ACC"
    assert wb["Summary Counts"].cell(1, 1).value == "ITEM"
    assert wb["Summary Counts"].cell(2, 1).value == "Assigned Tickets"


def test_create_parsed_report_workbook_splits_actioned_and_not_actioned_tickets(tmp_path):
    output = tmp_path / "parsed-report.xlsx"

    create_parsed_report_workbook(
        output_path=output,
        assignments=[
            AssignmentTicket(account="SFKL1", client_name="Done Client", contact="0711111111", location="Site A", status="Assigned"),
            AssignmentTicket(account="SFKL2", client_name="Waiting Client", contact="0722222222", location="Site B", status="Assigned"),
        ],
        field_activity_reports=[],
        completions=[
            CompletionTicket(account="SFKL1", client_name="Done Client", contact="0711111111", location="Site A", remarks="Connected")
        ],
        summary_counts={"Assigned Tickets": 2, "Tickets done": 1, "Teams Available": 0},
    )

    wb = load_workbook(output)
    assert wb["Actioned Tickets"].cell(2, 1).value == "SFKL1"
    assert wb["Actioned Tickets"].cell(2, 10).value == "Connected"
    assert wb["Not Actioned Tickets"].cell(2, 1).value == "SFKL2"
    assert wb["Not Actioned Tickets"].cell(2, 10).value is None
